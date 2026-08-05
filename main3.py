import argparse
import atexit
import os
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from xlsxwriter import Workbook as XlsxWriterWorkbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError, sync_playwright
from greencheck_adapter import build_payloads
from greencheck_client import GreenCheckClient, GreenCheckError
from greencheck_queue import OutboundQueue
from greencheck_sources import load_sources
from greencheck_health import HealthReporter

BASE = Path(__file__).resolve().parent
STATE_FILE = BASE / "facebook_state.json"
GROUPS_FILE = BASE / "facebook_groups.txt"
XLSX_FILE = BASE / "facebook_posts_full.xlsx"
KEYWORDS_FILE = BASE / "keywords.txt"
GREENCHECK_QUEUE_FILE = BASE / "greencheck_outbound_queue.sqlite3"
GREENCHECK_CONFIG_CACHE = BASE / "greencheck_config_cache.json"

POSTS_SHEET = "Posts"
COMMENTS_SHEET = "Comments"
POST_SELECTOR = "div[data-ad-rendering-role='story_message']"
POST_ACTION_SELECTOR = "[aria-label^='Actions for this post by ']"
COMMENT_SELECTOR = (
    "div[role='article'][aria-label^='Comment by '], "
    "div[role='article'][aria-label^='Reply by ']"
)

# Small overlapping scrolls are deliberate. Facebook virtualizes the feed,
# so large jumps can move short posts out of the DOM before their permalink
# has finished hydrating.
SCROLL_PIXELS = 160
WAIT_MS = 550
MAX_SCROLLS = 160
STOP_AFTER_NO_NEW = 30
FIRST_RUN_POST_LIMIT = 5
INCREMENTAL_POST_LIMIT = 50
COMMENTS_PER_POST_LIMIT = 30
KNOWN_POSTS_TO_CONFIRM = 3
MAX_COMMENT_LOAD_ROUNDS = 20
COMMENT_WAIT_MS = 600
SOURCE_TIMEOUT_SECONDS = 180

POST_FIELDS = (
    "group_id", "group_name", "post_id", "post_datetime", "post_text",
    "displayed_poster", "poster_url", "post_url",
)
COMMENT_FIELDS = (
    "group_id", "group_name", "post_id", "comment_datetime",
    "comment_text", "commenter_name", "commenter_id",
)

INVALID_GROUP_NAMES = {
    "", "facebook", "groups", "facebook groups",
    "log into facebook", "notifications",
}

SHEET_SPECS = {
    POSTS_SHEET: {
        "fields": POST_FIELDS,
        "date_field": "post_datetime",
        "date_column": 4,
        "widths": (19, 30, 20, 20, 70, 25, 46, 52),
        "wrap": (5, 7, 8),
        "ids": (1, 3),
        "urls": (7, 8),
        "height": 60,
        "color": "1F4E78",
        "table": "FacebookPostsTable",
    },
    COMMENTS_SHEET: {
        "fields": COMMENT_FIELDS,
        "date_field": "comment_datetime",
        "date_column": 4,
        "widths": (19, 30, 20, 20, 80, 28, 20),
        "wrap": (5,),
        "ids": (1, 3, 7),
        "urls": (),
        "height": 45,
        "color": "548235",
        "table": "FacebookCommentsTable",
    },
}

ACTION_SCAN_JS = r"""
({gid, actionSelector, messageSelector}) => {
  const clean = value => (value || '').replace(/\s+/g, ' ').trim();

  const isRendered = element => {
    if (!element?.isConnected) return false;
    if (element.closest('[hidden], [aria-hidden="true"], [role="dialog"], [aria-modal="true"]')) return false;
    const style = getComputedStyle(element);
    if (style.display === 'none' || style.visibility === 'hidden') return false;
    const rect = element.getBoundingClientRect();
    return rect.width > 0 && rect.height > 0;
  };

  const before = (a, b) => Boolean(
    a.compareDocumentPosition(b) & Node.DOCUMENT_POSITION_FOLLOWING
  );

  const postRefFromHref = (href, scopes) => {
    if (!href) return null;
    const value = String(href).replace(/&amp;/g, '&');
    for (const scope of scopes) {
      const patterns = [
        new RegExp(`/groups/${scope}/(?:posts|permalink)/(\\d+)`),
        new RegExp(`/groups/${scope}/(?:posts|permalink)%2F(\\d+)`, 'i'),
      ];
      for (const pattern of patterns) {
        const match = value.match(pattern);
        if (match) return {id: match[1], scope};
      }
    }
    for (const pattern of [
      /[?&](?:multi_permalinks|story_fbid|fbid|post_id)=(\d+)/,
      /[?&]set=(?:pcb|gm)\.(\d+)/,
    ]) {
      const match = value.match(pattern);
      if (match) return {id: match[1], scope: gid};
    }
    return null;
  };

  const allActions = [...document.querySelectorAll(actionSelector)]
    .filter(isRendered);
  if (!allActions.length) return [];

  // Stay inside the actual group timeline. Prefer a rendered feed containing
  // links back to this group and crossing the browser's center line.
  const grouped = new Map();
  for (const action of allActions) {
    const feed = action.closest('[role="feed"]');
    if (!feed || !isRendered(feed)) continue;
    if (!grouped.has(feed)) grouped.set(feed, []);
    grouped.get(feed).push(action);
  }
  if (!grouped.size) return [];

  const centerX = innerWidth / 2;
  const ranked = [...grouped.entries()].map(([feed, actions]) => {
    const rect = feed.getBoundingClientRect();
    const groupLinks = [...feed.querySelectorAll('a[href]')]
      .filter(isRendered)
      .filter(link => String(link.href || '').includes(`/groups/${gid}/`)).length;
    const containsCenter = rect.left <= centerX && rect.right >= centerX;
    const actionTop = Math.min(...actions.map(a => a.getBoundingClientRect().top));
    return {feed, actions, groupLinks, containsCenter, actionTop, rect};
  }).sort((a, b) =>
    Number(b.groupLinks > 0) - Number(a.groupLinks > 0) ||
    Number(b.containsCenter) - Number(a.containsCenter) ||
    b.groupLinks - a.groupLinks ||
    Math.abs((a.rect.left + a.rect.right) / 2 - centerX) -
      Math.abs((b.rect.left + b.rect.right) / 2 - centerX) ||
    a.actionTop - b.actionTop
  );

  const feed = ranked[0].feed;
  const feedActions = ranked[0].actions;
  // Vanity group URLs (for example /groups/nocateehomes/) are rendered by
  // Facebook with a numeric scope in member and permalink URLs. Learn only
  // from group-member links in this selected feed, and prefer the scope seen
  // most often. The configured opaque group ID remains unchanged in storage.
  const scopeCounts = new Map();
  for (const link of feed.querySelectorAll('a[href]')) {
    const match = String(link.href || '').match(/\/groups\/([^/?#]+)\/user\/\d+/);
    if (match) scopeCounts.set(match[1], (scopeCounts.get(match[1]) || 0) + 1);
  }
  const renderedScope = [...scopeCounts.entries()]
    .sort((a, b) => b[1] - a[1] || a[0].localeCompare(b[0]))[0]?.[0];
  const scopes = [...new Set([gid, renderedScope].filter(Boolean))];
  const counter = window.__fbScraperActionCounter ||= {value: 0};
  const timeCounter = window.__fbScraperTimeCounter ||= {value: 0};

  // A normal Facebook feed child is one feed item. Taking only the first post
  // action in each child prevents a nested shared post from becoming a second
  // timeline item.
  const units = new Map();
  for (const action of feedActions) {
    let unit = action;
    while (unit.parentElement && unit.parentElement !== feed) {
      unit = unit.parentElement;
    }
    if (!units.has(unit)) units.set(unit, []);
    units.get(unit).push(action);
  }

  const results = [];
  for (const [unit, actions] of units.entries()) {
    actions.sort((a, b) =>
      a.getBoundingClientRect().top - b.getBoundingClientRect().top
    );
    const action = actions[0];
    if (!action) continue;

    if (!action.dataset.fbScraperActionId) {
      counter.value += 1;
      action.dataset.fbScraperActionId = String(counter.value);
    }

    const links = [...unit.querySelectorAll('a[href]')].filter(isRendered);
    const headerLinks = links.filter(link => before(link, action));
    const candidates = [];
    const addId = value => {
      const ref = postRefFromHref(value, scopes);
      if (ref && !candidates.some(item => item.id === ref.id)) candidates.push(ref);
    };

    // The outer post's timestamp/permalink is in the header, before its menu.
    for (const link of headerLinks) addId(link.href);

    // A visible comment permalink still contains the owning outer post ID.
    if (!candidates.length) {
      for (const link of links) {
        if (/[?&]comment_id=\d+/.test(link.href || '')) addId(link.href);
      }
    }

    // Then accept one unique group-post ID from the whole feed item.
    if (!candidates.length) {
      const refs = links.map(link => postRefFromHref(link.href, scopes)).filter(Boolean);
      const ids = [...new Set(refs.map(ref => ref.id))];
      if (ids.length === 1) candidates.push(refs.find(ref => ref.id === ids[0]));
    }

    // Facebook sometimes keeps the same URL only in serialized DOM attributes.
    if (!candidates.length) {
      const html = unit.outerHTML
        .replace(/&quot;/g, '"')
        .replace(/\\u0025/gi, '%')
        .replace(/\\\//g, '/');
      const refs = [];
      for (const scope of scopes) {
        const pattern = new RegExp(`/groups/${scope}/(?:posts|permalink)/(\\d+)`, 'g');
        let match;
        while ((match = pattern.exec(html)) !== null) refs.push({id: match[1], scope});
      }
      const generic = /(?:story_fbid|top_level_post_id|post_id)["'=:\\s]+(?:"|&quot;)?(\d+)/g;
      let genericMatch;
      while ((genericMatch = generic.exec(html)) !== null) {
        refs.push({id: genericMatch[1], scope: gid});
      }
      const ids = [...new Set(refs.map(ref => ref.id))];
      if (ids.length === 1) candidates.push(refs.find(ref => ref.id === ids[0]));
    }

    const directRef = candidates[0] || null;
    const directId = directRef?.id || '';
    const directUrl = directId
      ? `https://www.facebook.com/groups/${directRef.scope}/posts/${directId}/`
      : '';

    const poster = clean(action.getAttribute('aria-label'))
      .replace('Actions for this post by ', '').trim();

    const messages = [...unit.querySelectorAll(messageSelector)]
      .filter(isRendered)
      .sort((a, b) => a.getBoundingClientRect().top - b.getBoundingClientRect().top);
    const message = messages[0] || null;
    const text = clean(message?.innerText) || '[Post without text]';

    let posterUrl = '';
    if (!poster.toLowerCase().startsWith('anonymous')) {
      const author = headerLinks.find(link =>
        clean(link.getAttribute('aria-label') || link.innerText) === poster
      );
      if (author) {
        const url = new URL(author.href, location.href);
        const member = url.pathname.match(/^\/groups\/([^/?#]+)\/user\/(\d+)/);
        if (member) {
          posterUrl = `https://www.facebook.com/groups/${member[1]}/user/${member[2]}/`;
        } else if (url.pathname === '/profile.php') {
          const id = url.searchParams.get('id');
          if (id) posterUrl = `https://www.facebook.com/profile.php?id=${id}`;
        } else {
          posterUrl = `${url.origin}${url.pathname}`;
        }
      }
    }

    const timestampLabel = link => clean(
      link.getAttribute('aria-label') || link.getAttribute('title') ||
      link.querySelector('[aria-label]')?.getAttribute('aria-label') ||
      link.innerText
    );
    const isTimestamp = link => {
      const label = timestampLabel(link);
      return link.target === '_blank' ||
        /\b(?:just now|\d+\s*(?:s|m|h|d|w|y)|today|yesterday)\b/i.test(label) ||
        /\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\b/i.test(label);
    };
    // Most cards put their timestamp before the menu in DOM order.  AI-content
    // cards can put it after the menu, so also accept a timestamp on the same
    // visual header line.  This excludes timestamps from a nested shared post.
    const actionTop = action.getBoundingClientRect().top;
    const timeLink = [...headerLinks].reverse().find(isTimestamp) ||
      links.filter(link => isTimestamp(link) &&
        !/[?&]comment_id=\d+/.test(link.href || '') &&
        link.getBoundingClientRect().top <= actionTop + 90
      ).sort((a, b) => b.getBoundingClientRect().top - a.getBoundingClientRect().top)[0];
    const postTime = timeLink ? clean(
      timeLink.getAttribute('aria-label') || timeLink.getAttribute('title') ||
      timeLink.querySelector('[aria-label]')?.getAttribute('aria-label') ||
      timeLink.innerText
    ) : '';

    let timeMarker = '';
    if (timeLink) {
      if (!timeLink.dataset.fbScraperTimeId) {
        timeCounter.value += 1;
        timeLink.dataset.fbScraperTimeId = String(timeCounter.value);
      }
      timeMarker = timeLink.dataset.fbScraperTimeId;
    }

    const hasGroupLink = links.some(link =>
      scopes.some(scope => String(link.href || '').includes(`/groups/${scope}/`))
    );
    const unitText = clean(unit.innerText).slice(0, 500);
    const sponsored = /(?:^|\\s)Sponsored(?:\\s|$)/i.test(unitText);
    const isGroupPost = Boolean(directId || hasGroupLink ||
      poster.toLowerCase().startsWith('anonymous')) &&
      !(sponsored && !directId && !hasGroupLink);

    const rect = action.getBoundingClientRect();
    results.push({
      marker: action.dataset.fbScraperActionId,
      time_marker: timeMarker,
      direct_url: directUrl,
      text,
      poster,
      poster_url: posterUrl,
      post_time: postTime,
      is_group_post: isGroupPost,
      absolute_top: window.scrollY + rect.top,
    });
  }

  return results.sort((a, b) => a.absolute_top - b.absolute_top);
}
"""

POST_PAGE_URL_JS = r"""
gid => {
  const values = [location.href];
  const canonical = document.querySelector('link[rel="canonical"]')?.href;
  const ogUrl = document.querySelector('meta[property="og:url"]')?.content;
  if (canonical) values.push(canonical);
  if (ogUrl) values.push(ogUrl);
  // When a timestamp opens an in-page post dialog, inspect only that visible
  // dialog. Scanning the underlying feed could select a neighboring post.
  for (const dialog of document.querySelectorAll('[role="dialog"]')) {
    const rect = dialog.getBoundingClientRect();
    if (!rect.width || !rect.height) continue;
    for (const link of dialog.querySelectorAll('a[href]')) {
      values.push(link.href || link.getAttribute('href') || '');
    }
  }

  const patterns = [
    /\/groups\/([^/?#]+)\/(?:posts|permalink)\/(\d+)/,
    /[?&](?:multi_permalinks|story_fbid|fbid|post_id)=(\d+)/,
    /[?&]set=(?:pcb|gm)\.(\d+)/,
  ];
  for (const value of values) {
    for (const pattern of patterns) {
      const match = String(value || '').replace(/&amp;/g, '&').match(pattern);
      if (match) {
        const scoped = match.length > 2;
        const scope = scoped ? match[1] : gid;
        const postId = scoped ? match[2] : match[1];
        return `https://www.facebook.com/groups/${scope}/posts/${postId}/`;
      }
    }
  }
  return '';
}
"""

POST_TIME_MAP_JS = r"""
() => {
  const times = {};
  const collect = root => {
    const stack = [root];
    while (stack.length) {
      const value = stack.pop();
      if (!value || typeof value !== 'object') continue;
      if (!Array.isArray(value) &&
          value.post_id != null &&
          value.creation_time != null) {
        const postId = String(value.post_id);
        const creationTime = Number(value.creation_time);
        if (/^\d+$/.test(postId) && Number.isFinite(creationTime)) {
          times[postId] = creationTime;
        }
      }
      const children = Array.isArray(value) ? value : Object.values(value);
      for (const child of children) {
        if (child && typeof child === 'object') stack.push(child);
      }
    }
  };

  for (const script of document.querySelectorAll('script[type="application/json"]')) {
    const raw = script.textContent || '';
    if (!raw.includes('"post_id"') || !raw.includes('"creation_time"')) continue;
    try { collect(JSON.parse(raw)); } catch {}
  }
  return times;
}
"""

COMMENT_EXTRACT_JS = r"""
({gid, expectedPostId, selector}) => {
  const clean = v => (v || '').replace(/\s+/g, ' ').trim();
  return [...document.querySelectorAll(selector)].map(article => {
    const label = clean(article.getAttribute('aria-label'));
    const links = [...article.querySelectorAll('a[href]')];
    const author = links.find(a => {
      const href = a.href || '';
      return (/\/groups\/[^/?#]+\/user\/\d+/.test(href) || href.includes('/profile.php')) &&
        clean(a.innerText || a.getAttribute('aria-label'));
    });

    let commenterName = clean(author?.innerText || author?.getAttribute('aria-label'));
    if (!commenterName) {
      const match = label.match(/^(?:Comment|Reply) by (.+?)(?: to .+?'s comment)?(?: \d+ .+ ago)?$/);
      commenterName = clean(match?.[1]);
    }

    let commenterId = '';
    if (author) {
      let match = (author.href || '').match(/\/groups\/[^/?#]+\/user\/(\d+)/);
      if (!match) match = (author.href || '').match(/[?&]id=(\d+)/);
      commenterId = match?.[1] || '';
    }

    let text = clean(article.querySelector('span[lang]')?.innerText);
    if (!text) {
      text = [...article.querySelectorAll('div[dir="auto"]')]
        .map(e => clean(e.innerText))
        .find(v => v && v !== commenterName &&
          !/^(?:Like|Reply|Share)$/i.test(v) &&
          !/^\d+\s*(?:s|m|h|d|w|y)$/i.test(v)) || '';
    }

    const permalink = links.find(a => /[?&]comment_id=\d+/.test(a.href || ''));
    const postMatch = permalink?.href.match(
      /\/groups\/[^/?#]+\/(?:posts|permalink)\/(\d+)/
    );
    let commentTime = clean(
      permalink?.getAttribute('aria-label') || permalink?.getAttribute('title') ||
      permalink?.querySelector('[aria-label]')?.getAttribute('aria-label')
    );
    if (!commentTime) {
      const relative = label.match(
        /\b(?:just now|\d+\s+(?:seconds?|minutes?|hours?|days?|weeks?|years?)\s+ago)$/i
      );
      commentTime = clean(relative?.[0]);
    }

    return {
      post_id: postMatch?.[1] || expectedPostId || '',
      comment_time: commentTime, comment_text: text,
      commenter_name: commenterName, commenter_id: commenterId
    };
  }).filter(x => x.comment_text || x.commenter_name || x.commenter_id);
}
"""

GROUP_NAME_JS = r"""
gid => {
  const clean = v => (v || '').replace(/\s+/g, ' ').trim();
  const headings = [...document.querySelectorAll('h1')];
  let memberHeading = '';
  for (const heading of headings) {
    let parent = heading;
    for (let depth = 0; parent && depth < 12; depth++, parent = parent.parentElement) {
      if (parent.querySelector(`a[href^="/groups/${gid}/members"]`)) {
        memberHeading = clean(heading.innerText);
        break;
      }
    }
    if (memberHeading) break;
  }
  return {
    title: clean(document.title),
    memberHeading,
    headings: headings.map(h => clean(h.innerText)).filter(Boolean)
  };
}
"""


def clean_text(value):
    value = str(value or "").translate(str.maketrans({
        "’": "'", "‘": "'", "“": '"', "”": '"',
        "–": "-", "—": "-", "…": "...", "\u00a0": " ",
    }))
    value = unicodedata.normalize("NFKD", value)
    return " ".join(value.encode("ascii", "ignore").decode().split())


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def load_group_ids():
    if not GROUPS_FILE.exists():
        raise FileNotFoundError(f"Missing group list: {GROUPS_FILE}")
    ids = []
    for raw in GROUPS_FILE.read_text(encoding="utf-8-sig").splitlines():
        line = raw.split("#", 1)[0].strip()
        match = re.search(r"(?:facebook\.com/groups/)?(\d{5,})", line)
        if line and not match:
            print(f"Skipping invalid group line: {raw}")
        elif match and match.group(1) not in ids:
            ids.append(match.group(1))
    if not ids:
        raise ValueError(f"No valid group IDs in {GROUPS_FILE}")
    return ids


def load_keywords():
    """Load case-insensitive whole-word highlighting terms from keywords.txt."""
    if not KEYWORDS_FILE.exists():
        return []
    return sorted({
        clean_text(line).lower()
        for line in KEYWORDS_FILE.read_text(encoding="utf-8-sig").splitlines()
        if clean_text(line) and not line.lstrip().startswith("#")
    }, key=len, reverse=True)


def highlight_parts(value, keywords):
    """Split text into normal and keyword runs for XlsxWriter rich strings."""
    text = cell_text(value)
    if not text or not keywords:
        return None
    pattern = re.compile(r"(?<!\w)(?:" + "|".join(map(re.escape, keywords)) + r")(?!\w)", re.I)
    matches = list(pattern.finditer(text))
    if not matches:
        return None

    parts = []
    start = 0
    for match in matches:
        if start < match.start():
            parts.append(text[start:match.start()])
        parts.append(match.group())
        start = match.end()
    if start < len(text):
        parts.append(text[start:])
    return parts


def click_visible(locator, timeout=700):
    clicked = 0
    for index in range(locator.count() - 1, -1, -1):
        try:
            item = locator.nth(index)
            if item.is_visible():
                item.click(timeout=timeout)
                clicked += 1
        except Exception:
            pass
    return clicked


def expand_see_more(page):
    total = 0
    for _ in range(3):
        clicked = click_visible(page.get_by_text("See more", exact=True))
        total += clicked
        if not clicked:
            break
        page.wait_for_timeout(250)
    return total


def scroll_feed(page):
    # Keep a large overlap between scans. A post can initially render before
    # Facebook inserts its permalink; repeated nearby scans let it hydrate.
    page.evaluate("pixels => window.scrollBy(0, pixels)", SCROLL_PIXELS)
    page.wait_for_timeout(WAIT_MS)


def normalize_time_map(raw):
    times = {}
    for post_id, value in (raw or {}).items():
        try:
            times[str(post_id)] = int(value)
        except (TypeError, ValueError):
            pass
    return times



def scan_post_actions(page, group_id):
    return page.evaluate(ACTION_SCAN_JS, {
        "gid": group_id,
        "actionSelector": POST_ACTION_SELECTOR,
        "messageSelector": POST_SELECTOR,
    }) or []


def canonical_post_url(value, group_id):
    post_id = extract_post_id(value)
    if not post_id:
        return "", ""
    rendered_scope = re.search(
        r"facebook\.com/groups/([^/?#]+)/(?:posts|permalink)/\d+", value or ""
    )
    url_scope = rendered_scope.group(1) if rendered_scope else group_id
    return post_id, f"https://www.facebook.com/groups/{url_scope}/posts/{post_id}/"


def post_url_from_open_page(opened_page, group_id, timeout_ms=10_000):
    """Read the exact group-post URL after Facebook opens a timestamp link."""
    deadline = datetime.now().timestamp() + timeout_ms / 1000
    while datetime.now().timestamp() < deadline:
        post_id, post_url = canonical_post_url(opened_page.url, group_id)
        if post_id:
            return post_id, post_url
        try:
            candidate = opened_page.evaluate(POST_PAGE_URL_JS, group_id)
        except Exception:
            candidate = ""
        post_id, post_url = canonical_post_url(candidate, group_id)
        if post_id:
            return post_id, post_url
        opened_page.wait_for_timeout(200)
    return "", ""


def resolve_timestamp_link(page, marker, group_id):
    """Click the post's own timestamp and read the exact permalink it opens.

    Facebook often hides the post ID from the timestamp anchor's href, but the
    normal click still opens the correct post in a new tab. This is more reliable
    than guessing from neighboring links or depending on the Copy link menu.
    """
    if not marker:
        return "", ""

    selector = f'[data-fb-scraper-time-id="{marker}"]'
    link = page.locator(selector).first
    try:
        if not link.is_visible(timeout=1500):
            return "", ""
        link.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        return "", ""

    # A few Facebook builds expose the real permalink directly after hydration.
    try:
        post_id, post_url = canonical_post_url(
            link.get_attribute("href") or "", group_id
        )
        if post_id:
            return post_id, post_url
    except Exception:
        pass

    original_url = page.url
    original_scroll = page.evaluate("window.scrollY")
    popup = None
    try:
        # Most timestamp links navigate the current tab.  Do not spend six
        # seconds waiting for a popup for every card whose URL is still hidden.
        with page.expect_popup(timeout=1500) as popup_info:
            link.click(timeout=4000)
        popup = popup_info.value
        try:
            popup.wait_for_load_state("domcontentloaded", timeout=15_000)
        except Exception:
            pass
        return post_url_from_open_page(popup, group_id)
    except PlaywrightTimeoutError:
        # Facebook may update the current tab or open an in-page dialog after
        # the popup wait expires. Poll the clicked page before deciding that the
        # timestamp is unresolved.
        post_id, post_url = post_url_from_open_page(page, group_id, 8000)
        if post_id:
            try:
                page.go_back(wait_until="domcontentloaded", timeout=20_000)
                page.wait_for_timeout(700)
                page.evaluate("position => window.scrollTo(0, position)", original_scroll)
                page.wait_for_timeout(500)
            except Exception:
                pass
            return post_id, post_url
    except Exception:
        pass
    finally:
        if popup is not None:
            try:
                popup.close()
            except Exception:
                pass

    return "", ""


def resolve_action_copy_link(page, marker, group_id):
    """Resolve one exact feed card through its own Copy link menu action.

    Some Facebook card variants expose neither a permalink href nor a working
    timestamp. The card's action menu is still deterministic because `marker`
    is attached to that card's own post-action control. Opaque Facebook share
    URLs are followed in an isolated tab and accepted only after they resolve
    to a concrete group post URL.
    """
    if not marker:
        return "", ""

    action = page.locator(
        f'[data-fb-scraper-action-id="{marker}"]'
    ).first
    try:
        if not action.is_visible(timeout=1500):
            return "", ""
        action.scroll_into_view_if_needed(timeout=3000)
        page.evaluate("navigator.clipboard.writeText('')")
        action.click(timeout=4000)
        menu = page.locator('[role="menu"]:visible').last
        menu.wait_for(state="visible", timeout=3000)
        copy_item = menu.get_by_text("Copy link", exact=True).first
        if not copy_item.is_visible(timeout=2000):
            page.keyboard.press("Escape")
            return "", ""
        copy_item.click(timeout=3000)
    except Exception:
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return "", ""

    copied = ""
    for _ in range(20):
        try:
            copied = page.evaluate("navigator.clipboard.readText()") or ""
        except Exception:
            copied = ""
        if copied:
            break
        page.wait_for_timeout(100)

    post_id, post_url = canonical_post_url(copied, group_id)
    if post_id:
        return post_id, post_url
    if not re.match(r"^https://(?:www\.)?facebook\.com/", copied):
        return "", ""

    opened_page = page.context.new_page()
    try:
        opened_page.goto(copied, wait_until="domcontentloaded", timeout=30_000)
        return post_url_from_open_page(opened_page, group_id, 10_000)
    except Exception:
        return "", ""
    finally:
        try:
            opened_page.close()
        except Exception:
            pass


def extract_creation_times(page):
    return normalize_time_map(page.evaluate(POST_TIME_MAP_JS))


def select_all_comments(page):
    try:
        current = page.locator("[role='button']").filter(
            has_text=re.compile(r"^Most relevant", re.I)
        ).first
        if not current.is_visible(timeout=1000):
            return False
        current.click(timeout=1500)
        page.wait_for_timeout(300)
        option = page.locator(
            "[role='menuitem'], [role='menuitemradio'], "
            "[role='option'], [role='button']"
        ).filter(has_text=re.compile(r"^All comments", re.I)).first
        if option.is_visible(timeout=1000):
            option.click(timeout=1500)
            page.wait_for_timeout(500)
            return True
    except Exception:
        pass
    return False


def load_comments_up_to_limit(page, limit=COMMENTS_PER_POST_LIMIT):
    """Load comments/replies until the requested cap or the thread is exhausted."""
    select_all_comments(page)
    comments = page.locator(COMMENT_SELECTOR)
    load_pattern = re.compile(
        r"^(?:View|See|Load)\s+(?:(?:all|previous|more|\d+\s+more|\d+)\s+)?"
        r"(?:comments?|repl(?:y|ies))[\s\uFEFF]*$", re.I,
    )
    last_count, stable = -1, 0

    for _ in range(MAX_COMMENT_LOAD_ROUNDS):
        count = comments.count()
        if count >= limit:
            break

        clicked = click_visible(
            page.locator("[role='button']").filter(has_text=load_pattern), 1000
        )
        for index in range(comments.count() - 1, -1, -1):
            try:
                clicked += click_visible(
                    comments.nth(index).get_by_text("See more", exact=True)
                )
            except Exception:
                pass

        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(COMMENT_WAIT_MS)
        count = comments.count()
        stable = stable + 1 if count == last_count and clicked == 0 else 0
        last_count = count
        if stable >= 2:
            break

    # Expand truncated text for the comments that will actually be saved.
    for index in range(min(comments.count(), limit) - 1, -1, -1):
        try:
            click_visible(comments.nth(index).get_by_text("See more", exact=True))
        except Exception:
            pass
    page.wait_for_timeout(200)
    return comments.count()


def extract_comments(page, group_id, post_id):
    return page.evaluate(COMMENT_EXTRACT_JS, {
        "gid": group_id,
        "expectedPostId": post_id,
        "selector": COMMENT_SELECTOR,
    }) or []


def valid_group_name(value, group_id):
    value = clean_text(value)
    return "" if value.lower() in INVALID_GROUP_NAMES or value == group_id else value


def get_group_name(page, group_id):
    try:
        data = page.evaluate(GROUP_NAME_JS, group_id)
    except Exception:
        return group_id

    title = re.sub(r"^\(\d+\)\s*", "", clean_text(data.get("title")))
    title = re.sub(r"\s*[|·]\s*Facebook$", "", title, flags=re.I)
    for candidate in (title, data.get("memberHeading"), *(data.get("headings") or [])):
        name = valid_group_name(candidate, group_id)
        if name:
            return name
    return group_id


def as_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(value))
        except (ValueError, OverflowError):
            return None

    text = cell_text(value).strip()
    for pattern in ("%Y-%m-%d %H:%M:%S %z", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, pattern).replace(tzinfo=None)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def dt_text(value):
    parsed = as_datetime(value)
    return parsed.strftime("%Y-%m-%d %H:%M:%S") if parsed else ""


def facebook_datetime(value, now):
    value = clean_text(value).replace(" at ", " ")
    if not value:
        return ""

    if value.startswith("unix:"):
        try:
            parsed = datetime.fromtimestamp(int(value[5:]), tz=now.tzinfo)
            return parsed.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            return ""

    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if parsed.tzinfo:
            parsed = parsed.astimezone(now.tzinfo).replace(tzinfo=None)
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass

    lower = value.lower().strip(" .")
    parsed = None
    if lower in {"now", "just now"}:
        parsed = now
    elif lower.startswith(("today", "yesterday")):
        parsed = now - timedelta(days=lower.startswith("yesterday"))
        match = re.search(r"(\d{1,2}:\d{2}\s*[ap]m)", lower)
        if match:
            clock = datetime.strptime(match.group(1).upper(), "%I:%M %p")
            parsed = parsed.replace(hour=clock.hour, minute=clock.minute, second=0)
    else:
        verbose = re.fullmatch(
            r"(?:about\s+)?(\d+|an?|one)\s*"
            r"(second|minute|hour|day|week|year)s?(?:\s+ago)?", lower,
        )
        compact = re.fullmatch(r"(\d+)\s*([smhdwy])", lower)
        if verbose or compact:
            amount_text, unit = (verbose or compact).groups()
            amount = 1 if amount_text in {"a", "an", "one"} else int(amount_text)
            units = {
                "s": "seconds", "second": "seconds",
                "m": "minutes", "minute": "minutes",
                "h": "hours", "hour": "hours",
                "d": "days", "day": "days",
                "w": "weeks", "week": "weeks",
                "y": "days", "year": "days",
            }
            if unit in {"y", "year"}:
                amount *= 365
            parsed = now - timedelta(**{units[unit]: amount})

    if parsed:
        return parsed.replace(tzinfo=None, microsecond=0).strftime("%Y-%m-%d %H:%M:%S")

    for pattern in (
        "%A, %B %d, %Y %I:%M %p", "%B %d, %Y %I:%M %p",
        "%B %d %Y %I:%M %p", "%B %d, %Y",
    ):
        try:
            return datetime.strptime(value, pattern).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return ""


def post_key(row):
    group_id, post_id, url = (
        row.get("group_id", ""), row.get("post_id", ""), row.get("post_url", "")
    )
    if post_id:
        return f"id:{group_id}:{post_id}"
    if url:
        return f"url:{group_id}:{url}"
    return "fallback:" + "|".join((
        group_id, row.get("poster_url", ""), row.get("displayed_poster", ""),
        row.get("post_text", "")[:160],
    ))


def comment_key(row):
    return "|".join(row.get(field, "") for field in (
        "group_id", "post_id", "commenter_id", "commenter_name", "comment_text"
    ))


def extract_post_id(url):
    match = re.search(
        r"/(?:posts|permalink)/(\d+)|"
        r"[?&](?:multi_permalinks|story_fbid|fbid)=(\d+)|"
        r"[?&]set=(?:pcb|gm)\.(\d+)", url or ""
    )
    return next((value for value in match.groups() if value), "") if match else ""


def normalize_row(raw, fields, date_field):
    row = {field: cell_text(raw.get(field, "")) for field in fields}
    row[date_field] = dt_text(raw.get(date_field, ""))
    return row


def load_sheet(workbook, name, fields, date_field, key_function):
    if name not in workbook.sheetnames:
        return []

    values = workbook[name].iter_rows(values_only=True)
    header_row = next(values, None)
    if not header_row:
        return []

    headers = [cell_text(value) for value in header_row]
    rows, seen = [], set()
    for values_row in values:
        if not any(value not in (None, "") for value in values_row):
            continue
        raw = dict(zip(headers, values_row))
        row = normalize_row(raw, fields, date_field)
        key = key_function(row)
        if key not in seen:
            seen.add(key)
            rows.append(row)
    return rows


def load_existing_data():
    """Load existing post and comment rows from the workbook."""
    if not XLSX_FILE.exists():
        return [], []

    workbook = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    posts = load_sheet(
        workbook, POSTS_SHEET, POST_FIELDS, "post_datetime", post_key
    )
    comments = load_sheet(
        workbook, COMMENTS_SHEET, COMMENT_FIELDS, "comment_datetime", comment_key
    )
    workbook.close()
    return posts, comments


def merge_row(target, source, fields, long_field):
    for field in fields:
        incoming = source.get(field, "")
        if not incoming:
            continue
        if field == "group_name":
            target[field] = incoming
        elif field == long_field:
            if len(incoming) > len(target.get(field, "")):
                target[field] = incoming
        elif not target.get(field):
            target[field] = incoming


def merge_scan_order(order, scan_ids):
    """Merge one visible top-to-bottom scan into the accumulated feed order.

    Facebook virtualizes the feed, so consecutive scans overlap but may not
    contain exactly the same posts. This builds a shortest common supersequence
    of the existing order and the new scan, preserving the relative order from
    both while inserting posts that appear late.
    """
    scan = list(dict.fromkeys(cell_text(post_id) for post_id in scan_ids if post_id))
    if not scan:
        return
    if not order:
        order.extend(scan)
        return

    existing = list(dict.fromkeys(cell_text(post_id) for post_id in order if post_id))

    # With no overlap, the scan came from farther down the feed, so append it.
    if not set(existing).intersection(scan):
        order[:] = existing + [post_id for post_id in scan if post_id not in existing]
        return

    # LCS table used to create a shortest common supersequence.
    rows, columns = len(existing), len(scan)
    lcs = [[0] * (columns + 1) for _ in range(rows + 1)]
    for i in range(rows - 1, -1, -1):
        for j in range(columns - 1, -1, -1):
            if existing[i] == scan[j]:
                lcs[i][j] = 1 + lcs[i + 1][j + 1]
            else:
                lcs[i][j] = max(lcs[i + 1][j], lcs[i][j + 1])

    merged = []
    i = j = 0
    while i < rows and j < columns:
        if existing[i] == scan[j]:
            merged.append(existing[i])
            i += 1
            j += 1
        elif lcs[i + 1][j] >= lcs[i][j + 1]:
            merged.append(existing[i])
            i += 1
        else:
            merged.append(scan[j])
            j += 1
    merged.extend(existing[i:])
    merged.extend(scan[j:])

    order[:] = list(dict.fromkeys(merged))


def write_sheet(workbook, title, rows, spec, keywords):
    sheet = workbook.add_worksheet(title)
    fields = spec["fields"]
    header = workbook.add_format({"bold": True, "font_color": "FFFFFF", "bg_color": spec["color"], "align": "center", "valign": "vcenter", "text_wrap": True})
    plain = workbook.add_format({"valign": "top"})
    wrapped = workbook.add_format({"valign": "top", "text_wrap": True})
    date = workbook.add_format({"valign": "top", "num_format": "m/d/yy h:mm AM/PM"})
    link = workbook.add_format({"font_color": "0563C1", "underline": 1, "valign": "top", "text_wrap": True})
    red_bold = workbook.add_format({"bold": True, "font_color": "FF0000"})
    sheet.freeze_panes(1, 0)
    sheet.set_row(0, 26)
    for column, width in enumerate(spec["widths"]):
        sheet.set_column(column, column, width)
        sheet.write(0, column, fields[column], header)
    for row_number, row in enumerate(rows, 1):
        sheet.set_row(row_number, spec["height"])
        for column, field in enumerate(fields):
            value = row.get(field, "")
            if field == spec["date_field"] and as_datetime(value):
                sheet.write_datetime(row_number, column, as_datetime(value), date)
            elif column + 1 in spec["urls"] and str(value).startswith(("http://", "https://")):
                sheet.write_url(row_number, column, value, link, value)
            elif field in {"post_text", "comment_text"} and (parts := highlight_parts(value, keywords)):
                rich = []
                for index, part in enumerate(parts):
                    rich.extend((red_bold, part) if index % 2 else (part,))
                if len(rich) == 2:
                    sheet.write(row_number, column, value, red_bold)
                else:
                    sheet.write_rich_string(row_number, column, *rich, wrapped)
            else:
                sheet.write(row_number, column, cell_text(value), wrapped if column + 1 in spec["wrap"] else plain)
    if rows:
        sheet.add_table(0, 0, len(rows), len(fields) - 1, {"name": spec["table"], "style": "Table Style Medium 2", "columns": [{"header": field} for field in fields]})


def write_workbook(posts, comments):
    keywords = load_keywords()
    temporary = XLSX_FILE.with_name(f"{XLSX_FILE.stem}.tmp.xlsx")
    temporary.unlink(missing_ok=True)
    with XlsxWriterWorkbook(temporary) as workbook:
        write_sheet(workbook, POSTS_SHEET, posts, SHEET_SPECS[POSTS_SHEET], keywords)
        write_sheet(workbook, COMMENTS_SHEET, comments, SHEET_SPECS[COMMENTS_SHEET], keywords)
    temporary.replace(XLSX_FILE)


def scrape_post_comments(page, post):
    if not post.get("post_url"):
        return []
    page.goto(post["post_url"], wait_until="domcontentloaded", timeout=60_000)
    page.wait_for_timeout(1000)
    now = datetime.now().astimezone()

    exact = extract_creation_times(page).get(str(post.get("post_id", "")))
    if exact:
        post["post_datetime"] = facebook_datetime(f"unix:{exact}", now)

    rendered = load_comments_up_to_limit(page, COMMENTS_PER_POST_LIMIT)
    rows, seen = [], set()
    for raw in extract_comments(page, post["group_id"], post["post_id"]):
        text = clean_text(raw.get("comment_text"))
        if not text:
            continue

        row = {
            "group_id": post["group_id"],
            "group_name": post["group_name"],
            "post_id": raw.get("post_id") or post["post_id"],
            "comment_datetime": facebook_datetime(raw.get("comment_time"), now),
            "comment_text": text,
            "commenter_name": clean_text(raw.get("commenter_name")),
            "commenter_id": cell_text(raw.get("commenter_id")),
        }
        key = comment_key(row)
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
        if len(rows) >= COMMENTS_PER_POST_LIMIT:
            break

    print(
        f"Comments for post {post['post_id']}: saved {len(rows)} "
        f"(limit {COMMENTS_PER_POST_LIMIT}) from {rendered} rendered elements"
    )
    return rows



def scrape_group(page, source, known_keys, first_run, initial_post_limit):
    """Walk the visible feed in exact top-to-bottom post-action order.

    Each feed item is resolved from its own DOM permalink first; Facebook's
    Copy link command is only a fallback. A real unresolved post stops the group
    instead of being silently skipped.
    """
    group_id = source["group_id"]
    group_name = source["group_name"]
    source_url = source["group_url"]
    if source["facebook_source_type"] != "group":
        raise RuntimeError("Facebook page extraction is not implemented yet; source was not scraped.")
    scope = re.search(r"facebook\.com/groups/([^/?#]+)", source_url)
    if not scope:
        raise RuntimeError("Configured group URL does not contain a Facebook group path.")
    facebook_scope = scope.group(1)
    page.goto(source_url, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.locator(POST_ACTION_SELECTOR).first.wait_for(timeout=30_000)
    except PlaywrightTimeoutError:
        join_controls = page.get_by_text("Join group", exact=True)
        visible_join = any(
            join_controls.nth(index).is_visible()
            for index in range(join_controls.count())
        )
        if visible_join:
            raise RuntimeError(
                "No group posts are visible and Facebook offers Join group; "
                "join approval may be required before posts can be scraped."
            ) from None
        raise
    page.evaluate("window.scrollTo(0, 0)")
    page.wait_for_timeout(1200)

    rows_by_id = {}
    order = []
    required_known = min(KNOWN_POSTS_TO_CONFIRM, len(known_keys))
    stable_signature = None
    stable_scans = 0
    no_order_change = 0
    unresolved_attempts = {}
    deadline = datetime.now().timestamp() + SOURCE_TIMEOUT_SECONDS

    def is_known(post_id):
        return f"id:{group_id}:{post_id}" in known_keys

    def boundary_start():
        if required_known <= 0:
            return None
        run = 0
        for index, post_id in enumerate(order):
            if is_known(post_id):
                run += 1
                if run >= required_known:
                    return index - required_known + 1
            else:
                run = 0
        return None

    for scan_number in range(1, MAX_SCROLLS + 1):
        if datetime.now().timestamp() >= deadline:
            raise RuntimeError(
                f"Source time limit of {SOURCE_TIMEOUT_SECONDS} seconds reached; "
                "no partial source data was saved."
            )
        # Text truncation is expanded before the DOM snapshot, so the workbook
        # receives the post body rather than Facebook's literal “See more”.
        expand_see_more(page)
        raw_actions = scan_post_actions(page, facebook_scope)
        if not raw_actions:
            page.wait_for_timeout(WAIT_MS)
            continue

        scan_ids = []
        now = datetime.now().astimezone()
        interactive_resolution_used = False
        retry_unresolved = False

        for raw in raw_actions:
            marker = cell_text(raw.get("marker"))
            if not marker:
                continue

            # Resolve from the exact feed item's own links first. This avoids
            # opening a menu for every post and avoids stale marker-to-ID caches
            # when Facebook recycles DOM nodes while scrolling.
            post_id, post_url = canonical_post_url(
                raw.get("direct_url", ""), facebook_scope
            )

            if not post_id and not raw.get("is_group_post", True):
                # Ads and non-post feed modules may expose a post-style menu but
                # are not actual posts from this group.
                continue

            if not post_id:
                post_id, post_url = resolve_timestamp_link(
                    page, cell_text(raw.get("time_marker")), facebook_scope
                )
                interactive_resolution_used = bool(post_id)

            if not post_id:
                post_id, post_url = resolve_action_copy_link(
                    page, marker, facebook_scope
                )
                interactive_resolution_used = bool(post_id)

            if not post_id:
                preview = clean_text(raw.get("text"))[:100]
                unresolved_key = "|".join((
                    clean_text(raw.get("poster")),
                    clean_text(raw.get("post_time")),
                    preview,
                ))
                attempts = unresolved_attempts.get(unresolved_key, 0) + 1
                unresolved_attempts[unresolved_key] = attempts
                if attempts <= 3:
                    retry_unresolved = True
                    print(
                        "Waiting for Facebook to hydrate one unresolved card "
                        f"({attempts}/3): {preview!r}"
                    )
                    break
                raise RuntimeError(
                    "A visible group post could not be opened through its "
                    f"timestamp: {preview!r}. Workbook was not changed."
                )

            row = {
                "group_id": group_id,
                "group_name": group_name,
                "post_id": post_id,
                "post_datetime": facebook_datetime(raw.get("post_time"), now),
                "post_text": clean_text(raw.get("text")) or "[Post without text]",
                "displayed_poster": clean_text(raw.get("poster")),
                "poster_url": raw.get("poster_url", ""),
                "post_url": post_url,
            }

            if post_id in rows_by_id:
                merge_row(rows_by_id[post_id], row, POST_FIELDS, "post_text")
            else:
                rows_by_id[post_id] = row
            scan_ids.append(post_id)

            # Returning from a timestamp permalink causes Facebook to rebuild
            # the virtualized feed. Markers captured in raw_actions are no
            # longer safe, so merge this resolved prefix and take a fresh DOM
            # snapshot before attempting another hidden permalink.
            if interactive_resolution_used:
                break

        scan_ids = list(dict.fromkeys(scan_ids))
        before_order = tuple(order)
        merge_scan_order(order, scan_ids)
        order_changed = tuple(order) != before_order
        no_order_change = 0 if order_changed else no_order_change + 1

        if retry_unresolved:
            page.wait_for_timeout(WAIT_MS)
            continue

        boundary = boundary_start() if not first_run else None
        if first_run:
            signature = tuple(order[:initial_post_limit]) if len(order) >= initial_post_limit else None
            collected = min(len(order), initial_post_limit)
            print(
                f"{group_name} [{group_id}] scan {scan_number}: "
                f"{collected}/{initial_post_limit} top posts found"
            )
        else:
            prefix = order[:boundary] if boundary is not None else order
            unseen_ids = [post_id for post_id in prefix if not is_known(post_id)]
            signature = tuple(unseen_ids) if boundary is not None else None
            boundary_text = "found" if boundary is not None else "not reached"
            print(
                f"{group_name} [{group_id}] scan {scan_number}: "
                f"{len(unseen_ids)} new posts; boundary {boundary_text}"
            )

        if signature is not None and signature == stable_signature and not order_changed:
            stable_scans += 1
        elif signature is not None:
            stable_signature = signature
            stable_scans = 1
        else:
            stable_signature = None
            stable_scans = 0

        if first_run and len(order) >= initial_post_limit and stable_scans >= 2:
            break

        if not first_run:
            if len(unseen_ids) >= INCREMENTAL_POST_LIMIT:
                print(
                    f"Incremental limit reached: keeping the first "
                    f"{INCREMENTAL_POST_LIMIT} new posts."
                )
                break
            if boundary is not None and stable_scans >= 2:
                print(f"Reached the stored-post boundary for {group_id}.")
                break

        if no_order_change >= STOP_AFTER_NO_NEW:
            break

        # Once a complete top window or boundary is present, rescan in place to
        # confirm it. Otherwise advance with heavy overlap.
        if (
            (first_run and len(order) >= initial_post_limit) or
            (not first_run and boundary is not None)
        ):
            page.wait_for_timeout(WAIT_MS)
        else:
            scroll_feed(page)

    if first_run:
        if len(order) < initial_post_limit:
            raise RuntimeError(
                f"Only {len(order)} posts were resolved; "
                f"need {initial_post_limit}. Workbook was not changed."
            )
        selected_ids = order[:initial_post_limit]
        rows = [rows_by_id[post_id] for post_id in selected_ids]
        print(
            f"First run for {group_id}: saving the first "
            f"{len(rows)} posts from the top."
        )
        for index, row in enumerate(rows, 1):
            preview = row.get("post_text", "")[:80]
            print(f"  {index}. {row['post_id']} | {preview}")
    else:
        boundary = boundary_start()
        prefix = order[:boundary] if boundary is not None else order
        unseen_ids = [post_id for post_id in prefix if not is_known(post_id)]
        if boundary is None and len(unseen_ids) < INCREMENTAL_POST_LIMIT:
            raise RuntimeError(
                "Stored-post boundary was not reached. Workbook was not changed."
            )
        selected_ids = unseen_ids[:INCREMENTAL_POST_LIMIT]
        rows = [rows_by_id[post_id] for post_id in selected_ids]
        print(
            f"Incremental run for {group_id}: adding {len(rows)} new posts "
            f"(limit {INCREMENTAL_POST_LIMIT})."
        )

    return rows, group_name

def sort_key(row):
    parsed = as_datetime(row.get("post_datetime") or row.get("comment_datetime"))
    timestamp = parsed.timestamp() if parsed else float("-inf")
    post_id = row.get("post_id", "")
    return timestamp, int(post_id) if str(post_id).isdigit() else -1


def main():
    parser = argparse.ArgumentParser(
        description="Scrape Facebook groups into an Excel workbook."
    )
    parser.add_argument(
        "--initial-post-limit", type=int, default=FIRST_RUN_POST_LIMIT,
        help="Posts to capture for a group with no existing workbook rows (default: 5).",
    )
    parser.add_argument("--offline-debug-groups", action="store_true", help="Use facebook_groups.txt only for local debugging; never use this in production.")
    args = parser.parse_args()
    if args.initial_post_limit < 1:
        parser.error("--initial-post-limit must be at least 1")

    if not STATE_FILE.exists():
        raise FileNotFoundError(f"Missing login state: {STATE_FILE}")

    if args.offline_debug_groups:
        sources = [{"group_id": value, "group_name": value, "group_url": f"https://www.facebook.com/groups/{value}", "facebook_source_type": "group", "scrape_enabled": True} for value in load_group_ids()]
        greencheck = queue = health = None
    else:
        greencheck = GreenCheckClient()
        sources, stale_config = load_sources(greencheck, GREENCHECK_CONFIG_CACHE)
        queue = OutboundQueue(GREENCHECK_QUEUE_FILE)
        health = HealthReporter(greencheck, queue, len(sources))
        health.start()
        atexit.register(health.stop)
        if stale_config:
            print("Using stale Green Check source configuration after a temporary API outage.")
    existing_posts, existing_comments = load_existing_data()
    posts = {post_key(row): row for row in existing_posts}
    comments = {comment_key(row): row for row in existing_comments}
    posts_by_group, comments_by_group = defaultdict(list), defaultdict(list)
    for row in posts.values():
        posts_by_group[row.get("group_id", "")].append(row)
    for row in comments.values():
        comments_by_group[row.get("group_id", "")].append(row)

    new_posts, new_comments, new_comment_rows = [], 0, []
    successful_groups = 0
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(storage_state=str(STATE_FILE))
        try:
            context.grant_permissions(
                ["clipboard-read", "clipboard-write"],
                origin="https://www.facebook.com",
            )
        except Exception:
            pass
        page = context.new_page()
        if health:
            page.goto("https://www.facebook.com/", wait_until="domcontentloaded", timeout=60_000)
            health.set_browser_status(
                "logged_out" if page.get_by_text("Log In", exact=True).count() else "healthy"
            )

        for source in sources:
            group_id = source["group_id"]
            if health:
                health.begin_source(source)
            if source["facebook_source_type"] != "group":
                reason = (
                    f"Unsupported Facebook source type: {source['facebook_source_type']}"
                )
                print(f"Skipping source {group_id}: {reason}")
                if health:
                    health.skip_source(source, reason)
                continue
            known_keys = {post_key(row) for row in posts_by_group[group_id]}
            first_run = not known_keys

            try:
                scraped, group_name = scrape_group(
                    page, source, known_keys, first_run, args.initial_post_limit
                )
            except Exception as error:
                print(f"Failed group {group_id}: {error}")
                if health:
                    health.fail_source(source, error)
                continue

            successful_groups += 1
            if health:
                health.complete_source(source, len(scraped))

            if group_name != group_id:
                for row in (*posts_by_group[group_id], *comments_by_group[group_id]):
                    row["group_name"] = group_name
            print(f"Using group name: {group_name} [{group_id}]")

            for row in scraped:
                key = post_key(row)
                if key in posts:
                    merge_row(posts[key], row, POST_FIELDS, "post_text")
                else:
                    posts[key] = row
                    posts_by_group[group_id].append(row)
                    new_posts.append(row)

        # Every post added during this execution is opened and gets up to 30
        # unique comments/replies. Existing posts are not re-scraped.
        posts_to_scrape = {
            post_key(row): row for row in new_posts if row.get("post_url")
        }

        for post in posts_to_scrape.values():
            try:
                scraped_comments = scrape_post_comments(page, post)
            except Exception as error:
                print(f"Failed comments for post {post.get('post_id', '')}: {error}")
                if health:
                    health.source_error(post["group_id"], error)
                continue

            if health:
                health.add_comments(post["group_id"], len(scraped_comments))

            for row in scraped_comments:
                key = comment_key(row)
                if key in comments:
                    merge_row(comments[key], row, COMMENT_FIELDS, "comment_text")
                else:
                    comments[key] = row
                    comments_by_group[row["group_id"]].append(row)
                    new_comments += 1
                    new_comment_rows.append(row)

        browser.close()

    if health:
        health.finish_cycle()

    if successful_groups == 0:
        print("No groups were scraped successfully. Workbook was not changed.")
        if health:
            health.stop()
            atexit.unregister(health.stop)
        return

    # Preserve Facebook feed order. Newly discovered posts are prepended in the
    # exact top-to-bottom order returned by scrape_group; existing spreadsheet
    # rows retain their previous order underneath them. Sorting by creation time
    # here would undo the feed order and could put a deeper post in row 1.
    new_post_keys = list(dict.fromkeys(post_key(row) for row in new_posts))
    new_post_key_set = set(new_post_keys)
    existing_order = [
        post_key(row) for row in existing_posts
        if post_key(row) not in new_post_key_set
    ]
    existing_order_set = set(existing_order)
    remaining_keys = [
        key for key in posts
        if key not in new_post_key_set and key not in existing_order_set
    ]
    final_post_keys = new_post_keys + existing_order + remaining_keys
    final_posts = [posts[key] for key in final_post_keys if key in posts]

    final_comments = sorted(comments.values(), key=sort_key, reverse=True)
    write_workbook(final_posts, final_comments)
    if greencheck and final_posts:
        # The server deduplicates stable Facebook IDs. Sending the local
        # workbook snapshot bootstraps an empty server safely, then keeps it
        # converged while local delivery state is being established.
        payloads = build_payloads(
            final_posts, final_comments, greencheck.client_id, "0.1.0", sources
        )
        for batch_id, payload in payloads:
            queue.enqueue(batch_id, payload)
        print(f"Queued {len(payloads)} stable Green Check batch(es).")
        for pending_id, raw, attempts in queue.pending():
            try:
                result = greencheck.ingest(__import__("json").loads(raw))
                if result.get("batch_id") == pending_id:
                    queue.delivered(pending_id)
            except GreenCheckError as error:
                queue.failure(pending_id, attempts, str(error), not error.temporary)
    if health:
        health.stop()
        atexit.unregister(health.stop)
    print(
        f"Inserted {len(new_posts)} new posts and {new_comments} new comments "
        f"into {XLSX_FILE.name}; {len(final_posts)} unique posts and "
        f"{len(final_comments)} unique comments total."
    )


if __name__ == "__main__":
    main()
